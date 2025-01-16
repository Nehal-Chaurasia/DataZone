import boto3
import pandas as pd
from botocore.exceptions import ClientError
import time
import os
from openpyxl import load_workbook
import sys

if not sys.warnoptions:
    import warnings
    warnings.simplefilter("ignore")

access_key = os.environ['AWS_ACCESS_KEY']
secret_key = os.environ['AWS_SECRET_KEY']
# session_token = os.environ["AWS_SESSION_TOKEN"]
region = 'eu-west-1'

domain_name = 'te-dip'
project_name = 'Sub-domains'
current_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(current_dir, "Collinson Business Glossary v1.1.xlsx")

sheet_name = 'Business Glossary'

# Create a Glossary.
def create_glossary(domain_name_id, glossary_name, description, project_id):
    response = datazone_client.create_glossary(
        domainIdentifier=domain_name_id,
        name=glossary_name,
        description=description,
        owningProjectIdentifier=project_id
    )
    
# Create a Business Term.
def create_business_term(domain_name_id, glossary_id, description, business_term):
    response = datazone_client.create_glossary_term(
        domainIdentifier=domain_name_id,
        glossaryIdentifier=glossary_id,
        longDescription=description,
        name=business_term,
        status='ENABLED'
    )

def create_glossary_list():
    response = datazone_client.search(domainIdentifier=domain_id.get(domain_name), searchScope='GLOSSARY')
    glossary_id = {}
    for glossary in response['items']:
        glossary_name = glossary['glossaryItem']['name']
        glossary_name_id = glossary['glossaryItem']['id']
        glossary_id[glossary_name] = glossary_name_id
    return glossary_id

def create_glossary_term_list():
    paginator = datazone_client.get_paginator('search')
    response = paginator.paginate(domainIdentifier=domain_id.get(domain_name), searchScope='GLOSSARY_TERM')
    glossary_term_list = {}
    for page in response:
        for glossary in page.get('items', []):
            glossary_term_name = glossary['glossaryTermItem']['name']
            glossary_term_id = glossary['glossaryTermItem']['id']
            glossary_id = glossary['glossaryTermItem']['glossaryId']
            glossary_term_list[glossary_term_name] = {'glossary_term_id': glossary_term_id, 'glossary_id':glossary_id}
    return glossary_term_list

# Create Boto3 client for AWS DataZone
datazone_client = boto3.client(
                    'datazone',
                    region_name = 'eu-west-1',
                    aws_access_key_id=access_key,
                    aws_secret_access_key=secret_key,
                    # aws_session_token = session_token,
                    verify=False)

# Create list of domain's
response = datazone_client.list_domains()
domain_id = {}
for domain in response['items']:
    domain_name = domain['name']
    domain_name_id = domain['id']
    domain_id[domain_name] = domain_name_id

# Create list of project's
response = datazone_client.list_projects(domainIdentifier=domain_id.get(domain_name))
project_id = {}
for project in response['items']:
    project_name = project['name']
    project_name_id = project['id']
    project_id[project_name] = project_name_id

workbook = load_workbook(excel_file_path)
sheet = workbook[sheet_name]

# Create a mapping of column names to their indices
header_mapping = {cell.value: idx + 1 for idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)))}
glossary_col = header_mapping["Glossary"]
business_term_col = header_mapping["Business Term"]
desc_col = header_mapping["Description"]
# action_col = header_mapping["Action"]

# Preprocess Data
rows_to_process = []
for row in range(2, sheet.max_row + 1):  # Skip header row
    glossary = sheet.cell(row=row, column=glossary_col).value
    business_term = sheet.cell(row=row, column=business_term_col).value
    description = sheet.cell(row=row, column=desc_col).value
    # action = sheet.cell(row=row, column=action_col).value

    # Filter rows and handle empty descriptions
    if glossary and business_term:
        # if action.lower() in ['create', 'update']:
        rows_to_process.append((row, glossary, business_term, description or "No Description"))

for row, glossary, business_term, description in rows_to_process:
    # if action.lower() == 'create':
    glossary_id = create_glossary_list()
    if not glossary.lower() in (key.lower() for key in glossary_id):
        create_glossary(domain_id.get(domain_name), glossary, 'No Description', project_id.get(project_name))
        time.sleep(30)

    glossary_id = create_glossary_list()
    if glossary not in glossary_id.keys():
        print("Sleeping for 2 mins for glossary to get created.")
        time.sleep(120)

    glossary_id = create_glossary_list()
    try:
        print(f"Creating business term {business_term}")
        create_business_term(domain_id.get(domain_name), glossary_id.get(glossary), description, business_term)
    except datazone_client.exceptions.ConflictException as e:
        print(f"Conflict: Glossary term '{business_term}' already exists. Skipping.")

    # sheet.cell(row=row, column=action_col, value="No")
    # workbook.save(excel_file_path)
    
    # elif action.lower() == 'update':
    #     try:
    #         print(f"Updating business term: {business_term}")
    #         glossary_term_list = create_glossary_term_list()
    #         glossary_term_name = glossary_term_list.get(business_term)
    #         response = datazone_client.update_glossary_term(
    #                 domainIdentifier=domain_id.get(domain_name),
    #                 glossaryIdentifier=glossary_id.get(glossary),
    #                 identifier=glossary_term_name.get('glossary_term_id'),
    #                 longDescription=description
    #             )
    #     except:
    #         print("Business Term is not created")
    #     sheet.cell(row=row, column=action_col, value="No")
    #     workbook.save(excel_file_path)
    